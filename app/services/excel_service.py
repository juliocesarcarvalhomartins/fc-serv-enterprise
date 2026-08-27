from __future__ import annotations

import re
import shutil
import unicodedata
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..config import BACKUP_DIR
from ..models import Invoice
from .invoice_rules import (
    GKO_OK_VALUES,
    SAVE_OK_VALUES,
    carrier_key,
    normalize_invoice_number,
    parse_amount,
    parse_date,
    status_is_ok,
)


def _norm(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "").strip())
    text = "".join(c for c in text if not unicodedata.combining(c)).upper()
    return re.sub(r"[^A-Z0-9]+", " ", text).strip()


HEADER_ALIASES = {
    "carrier": {"TRANSPORTADORA", "FORNECEDOR", "FORNECEDOR TRANSPORTADORA", "PRESTADOR", "EMPRESA", "RAZAO SOCIAL"},
    "number": {"FATURA", "NUMERO FATURA", "N FATURA", "NO FATURA", "NF", "NFS E", "NOTA FISCAL", "NUMERO", "DOCUMENTO"},
    "inclusion": {"DATA INCLUSAO", "INCLUSAO", "DATA CADASTRO", "DATA RECEBIMENTO", "DATA"},
    "due": {"VENCIMENTO", "DATA VENCIMENTO", "DT VENCIMENTO", "VCTO", "VENC"},
    "amount": {"VALOR", "VALOR TOTAL", "TOTAL", "VALOR FATURA", "VALOR DOCUMENTO"},
    "gko": {"SITUACAO GKO", "STATUS GKO", "GKO", "LIBERACAO GKO"},
    "save": {"STATUS FINAL SAVE", "STATUS SAVE", "SAVE", "SITUACAO SAVE", "FINAL SAVE"},
    "notes": {"OBSERVACOES", "OBSERVACAO", "NOTAS", "COMENTARIOS", "DESCRICAO", "SERVICO"},
    "type": {"TIPO", "TIPO FATURA", "CATEGORIA", "CLASSIFICACAO", "NATUREZA"},
}


def _find_header(worksheet) -> tuple[int, dict[str, int]]:
    """Localiza cabeçalhos mesmo quando a planilha mudou de linha ou de nomes."""
    best_row = 0
    best_map: dict[str, int] = {}
    for row_number, row in enumerate(worksheet.iter_rows(min_row=1, max_row=min(40, worksheet.max_row), values_only=True), start=1):
        found: dict[str, int] = {}
        for index, value in enumerate(row):
            header = _norm(value)
            if not header:
                continue
            for field, aliases in HEADER_ALIASES.items():
                if header in aliases and field not in found:
                    found[field] = index
        # Fornecedores e números são indispensáveis; os demais são opcionais.
        score = len(found) + (3 if "carrier" in found else 0) + (3 if "number" in found else 0)
        current_score = len(best_map) + (3 if "carrier" in best_map else 0) + (3 if "number" in best_map else 0)
        if score > current_score:
            best_row, best_map = row_number, found
    if "carrier" in best_map and "number" in best_map:
        return best_row, best_map
    return 0, {}


def _choose_sheet(workbook):
    preferred = [
        "CONTROLE DE FATURAS", "FATURAS", "CONTROLE FATURAS", "CONTAS A PAGAR",
        "FRETES", "FRETES A PAGAR", "SERVICOS", "SERVIÇOS", "SERVICOS A PAGAR",
    ]
    by_norm = {_norm(name): workbook[name] for name in workbook.sheetnames}
    for name in preferred:
        if _norm(name) in by_norm:
            return by_norm[_norm(name)]
    # Escolhe a primeira aba visível com dados, em vez de falhar por causa do nome.
    for worksheet in workbook.worksheets:
        if worksheet.sheet_state == "visible" and worksheet.max_row > 1 and worksheet.max_column > 1:
            return worksheet
    raise ValueError("Nenhuma aba com dados foi encontrada na planilha.")


def _cell(values: tuple, mapping: dict[str, int], field: str, fallback: int | None = None):
    index = mapping.get(field, fallback)
    if index is None or index >= len(values):
        return None
    return values[index]


def _classify_excel_row(values: tuple, mapping: dict[str, int], sheet_title: str) -> tuple[str, int]:
    explicit = _norm(_cell(values, mapping, "type"))
    haystack = " ".join(_norm(value) for value in values if value not in (None, "")) + " " + _norm(sheet_title)
    service_terms = ("SERVICO", "SERVICOS", "NFS E", "NFSE", "PRESTACAO", "HONORARIO", "CONSULTORIA", "MANUTENCAO", "MENSALIDADE", "LICENCA", "ASSINATURA")
    freight_terms = ("FRETE", "CTE", "CT E", "CONHECIMENTO TRANSPORTE", "TRANSPORTADORA", "CARREGAMENTO", "ROMANEIO", "PEDAGIO")
    if explicit:
        if any(term in explicit for term in ("SERVICO", "NFS", "PRESTACAO")):
            return "SERVICE", 100
        if any(term in explicit for term in ("FRETE", "TRANSPORTE", "CTE")):
            return "FREIGHT", 100
    service_score = sum(1 for term in service_terms if term in haystack)
    freight_score = sum(1 for term in freight_terms if term in haystack)
    if service_score > freight_score and service_score:
        return "SERVICE", min(95, 65 + service_score * 8)
    # Planilhas antigas do FC SERV são de frete; preservar compatibilidade.
    return "FREIGHT", 90 if freight_score else 75


def import_workbook(db: Session, path: Path, user_id: int) -> dict:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    except (InvalidFileException, BadZipFile, OSError) as error:
        raise ValueError("O arquivo não é uma planilha XLSX válida ou está corrompido.") from error
    try:
        worksheet = _choose_sheet(workbook)
        header_row, mapping = _find_header(worksheet)
        # Compatibilidade com o modelo histórico: cabeçalho na linha 3 e dados na linha 4.
        if not mapping:
            header_row = 3
            mapping = {
                "carrier": 0, "number": 1, "inclusion": 2, "due": 4,
                "amount": 8, "gko": 9, "save": 11, "notes": 12,
            }

        inserted = duplicates = ignored = blank_rows = services = freights = 0
        existing = {
            (carrier, number)
            for carrier, number in db.execute(select(Invoice.carrier_key, Invoice.normalized_number)).all()
        }
        batch: set[tuple[str, str]] = set()
        start_row = max(1, header_row + 1)
        for row_number, row in enumerate(worksheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
            if row_number > 100_000:
                raise ValueError("A planilha ultrapassa o limite de 100.000 linhas de dados.")
            values = tuple(row)
            carrier = str(_cell(values, mapping, "carrier") or "").strip()
            number = str(_cell(values, mapping, "number") or "").strip()
            if not carrier and not number:
                blank_rows += 1
                ignored += 1
                if blank_rows >= 500:
                    break
                continue
            blank_rows = 0
            if not carrier or not number:
                ignored += 1
                continue
            key = (carrier_key(carrier), normalize_invoice_number(number))
            if not key[0] or not key[1]:
                ignored += 1
                continue
            if key in existing or key in batch:
                duplicates += 1
                continue

            gko_value = _cell(values, mapping, "gko")
            save_value = _cell(values, mapping, "save")
            save_posted = status_is_ok(save_value, SAVE_OK_VALUES)
            gko_released = save_posted or status_is_ok(gko_value, GKO_OK_VALUES)
            invoice_type, confidence = _classify_excel_row(values, mapping, worksheet.title)
            services += int(invoice_type == "SERVICE")
            freights += int(invoice_type == "FREIGHT")
            db.add(Invoice(
                carrier=carrier,
                carrier_key=key[0],
                invoice_number=number,
                normalized_number=key[1],
                inclusion_date=parse_date(_cell(values, mapping, "inclusion")) or date.today(),
                due_date=parse_date(_cell(values, mapping, "due")),
                amount=parse_amount(_cell(values, mapping, "amount")),
                situation_gko=str(gko_value or ""),
                gko_released=gko_released,
                save_posted=save_posted,
                status="COMPLETED" if save_posted else "PENDING",
                notes=str(_cell(values, mapping, "notes") or ""),
                source="excel",
                invoice_type=invoice_type,
                classification_confidence=confidence,
                created_by_id=user_id,
            ))
            batch.add(key)
            existing.add(key)
            inserted += 1
        db.commit()
        return {
            "inserted": inserted,
            "duplicates": duplicates,
            "ignored": ignored,
            "freights": freights,
            "services": services,
            "sheet": worksheet.title,
            "header_row": header_row,
        }
    finally:
        workbook.close()


def export_workbook(db: Session) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "CONTROLE DE FATURAS"
    headers = [
        "TRANSPORTADORA / FORNECEDOR", "FATURA", "DATA INCLUSÃO", "CHAVE", "VENCIMENTO",
        "COLUNA F", "COLUNA G", "COLUNA H", "VALOR", "SITUAÇÃO GKO",
        "COLUNA K", "STATUS FINAL SAVE", "OBSERVAÇÕES", "TIPO",
    ]
    worksheet.append(["FC SERV"])
    worksheet.append([f"Exportado em {datetime.now():%d/%m/%Y %H:%M}"])
    worksheet.append(headers)
    for cell in worksheet[3]:
        cell.fill = PatternFill("solid", fgColor="14345D")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    invoices = db.scalars(select(Invoice).order_by(Invoice.created_at.desc())).all()
    for invoice in invoices:
        worksheet.append([
            invoice.carrier, invoice.invoice_number, invoice.inclusion_date,
            f"{invoice.carrier_key}{invoice.normalized_number}", invoice.due_date,
            "", "", "", float(invoice.amount) if invoice.amount is not None else None,
            "OK" if invoice.gko_released else "PENDENTE", "",
            "OK" if invoice.save_posted else "PENDENTE", invoice.notes,
            "SERVIÇO" if invoice.invoice_type == "SERVICE" else "FRETE" if invoice.invoice_type == "FREIGHT" else "REVISAR",
        ])
    worksheet.freeze_panes = "A4"
    widths = [28, 16, 16, 30, 16, 12, 12, 12, 16, 18, 12, 20, 38, 16]
    from openpyxl.utils import get_column_letter
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    for row in worksheet.iter_rows(min_row=4):
        row[2].number_format = "dd/mm/yyyy"
        row[4].number_format = "dd/mm/yyyy"
        row[8].number_format = 'R$ #,##0.00'
        for cell in row:
            cell.alignment = Alignment(vertical="center")
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def backup_uploaded_file(path: Path) -> Path:
    target = BACKUP_DIR / f"{path.stem}-{datetime.now():%Y%m%d-%H%M%S}{path.suffix}"
    shutil.copy2(path, target)
    return target
